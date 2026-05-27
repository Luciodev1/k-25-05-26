from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db import models
from django.db.models import ProtectedError, Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.loader import render_to_string
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django_filters.views import FilterView
from io import BytesIO


class FinanceiroRequiredMixin(PermissionRequiredMixin):
    """Acesso restrito a utilizadores com permissoes financeiras."""
    permission_required = ('payments.add_payment',)

    def handle_no_permission(self):
        raise PermissionDenied("Nao tem permissao para aceder a esta pagina.")


class GestorRequiredMixin(PermissionRequiredMixin):
    """Acesso restrito a gestores (CRUD completo)."""
    permission_required = (
        'products.add_product',
        'products.change_product',
    )

    def handle_no_permission(self):
        raise PermissionDenied("Nao tem permissao para aceder a esta pagina.")


class AdminRequiredMixin(PermissionRequiredMixin):
    """Acesso restrito a administradores."""
    permission_required = ('auth.add_user', 'auth.change_user')

    def handle_no_permission(self):
        raise PermissionDenied("Nao tem permissao para aceder a esta pagina.")


class HtmxMixin:
    """Mixin para suportar requests HTMX com partial rendering.
    Define htmx_template_name no View para o template parcial.
    """
    htmx_template_name = None

    def get_template_names(self):
        if self.request.headers.get('HX-Request') and self.htmx_template_name:
            return [self.htmx_template_name]
        return super().get_template_names()


class ExportMixin:
    """Mixin para adicionar exportacao Excel/PDF a ListViews.
    Define export_columns como lista de (header, field_name) no View.
    """
    export_columns = []

    def get(self, request, *args, **kwargs):
        export = request.GET.get('export')
        if export in ('excel', 'pdf'):
            queryset = self.filterset.qs if hasattr(self, 'filterset') else self.get_queryset()
            if export == 'excel':
                return self._export_excel(queryset)
            else:
                return self._export_pdf(queryset)
        return super().get(request, *args, **kwargs)

    def _export_excel(self, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Headers
        for col_idx, (header, _) in enumerate(self.export_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, (_, field) in enumerate(self.export_columns, 1):
                value = obj
                for attr in field.split('.'):
                    value = getattr(value, attr, '')
                if hasattr(value, 'all'):
                    value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                cell.border = thin_border

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="exportacao.xlsx"'
        wb.save(response)
        return response

    def _export_pdf(self, queryset):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=1, spaceAfter=12)
        elements.append(Paragraph(f"Exportacao - {self.model._meta.verbose_name_plural.title()}", title_style))
        elements.append(Spacer(1, 0.5*cm))

        # Table data
        headers = [h for h, _ in self.export_columns]
        data = [headers]
        for obj in queryset:
            row = []
            for _, field in self.export_columns:
                value = obj
                for attr in field.split('.'):
                    value = getattr(value, attr, '')
                row.append(str(value) if value else '')
            data.append(row)

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="exportacao.pdf"'
        return response


# ── Tenant-aware mixin ─────────────────────────────────────────────
# Adiciona filtragem por tenant automaticamente a todas as views.


class TenantFilterMixin:
    tenant_field = 'tenant'

    def get_queryset(self):
        qs = super().get_queryset()
        tenant = getattr(self.request, 'tenant', None)
        if tenant:
            return qs.filter(**{self.tenant_field: tenant})
        return qs

    def form_valid(self, form):
        tenant = getattr(self.request, 'tenant', None)
        if tenant and hasattr(form.instance, self.tenant_field):
            setattr(form.instance, self.tenant_field, tenant)
        return super().form_valid(form)


class TenantCreateMixin(TenantFilterMixin):
    def get_initial(self):
        initial = super().get_initial()
        tenant = getattr(self.request, 'tenant', None)
        if tenant and hasattr(self.model, self.tenant_field):
            initial[self.tenant_field] = tenant
        return initial


# ── Generic CRUD base views ──────────────────────────────────────────
# Reduzem duplicação de código nos apps.
# Cada app define apenas model, form_class, permission_required, etc.


class BaseListView(LoginRequiredMixin, PermissionRequiredMixin, HtmxMixin, ExportMixin, TenantFilterMixin, FilterView):
    """ListView genérica com auth, permissões, HTMX, filtros, exportação e tenant."""
    paginate_by = 10


class BaseCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, TenantCreateMixin, CreateView):
    """CreateView genérica com auth, permissões, tenant e mensagem de sucesso."""
    pass


class BaseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, TenantFilterMixin, UpdateView):
    """UpdateView genérica com auth, permissões, tenant e mensagem de sucesso."""
    pass


class BaseDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantFilterMixin, DetailView):
    """DetailView genérica com auth, permissões e tenant."""
    pass


class BaseDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantFilterMixin, DeleteView):
    """DeleteView genérica com tratamento de ProtectedError e tenant."""
    protected_error_message = (
        'Nao e possivel eliminar este registo porque esta a ser utilizado '
        'por outros registos.'
    )

    def post(self, request, *args, **kwargs):
        try:
            obj = self.get_object()
            obj.delete()
            messages.success(request, self.success_message)
            return redirect(self.get_success_url())
        except ProtectedError:
            messages.error(request, self.protected_error_message)
            return redirect(self.get_success_url())


class BaseTrashListView(LoginRequiredMixin, PermissionRequiredMixin, HtmxMixin, TenantFilterMixin, ListView):
    """ListView para itens na lixeira (soft-deleted) com tenant."""
    paginate_by = 10

    def get_queryset(self):
        qs = self.model.all_objects.filter(is_deleted=True)
        tenant = getattr(self.request, 'tenant', None)
        if tenant:
            qs = qs.filter(**{self.tenant_field: tenant})
        return qs


class BaseRestoreView(LoginRequiredMixin, PermissionRequiredMixin, TenantFilterMixin, View):
    """View genérica para restaurar um item da lixeira com tenant."""
    model = None
    redirect_url = None

    def post(self, request, pk):
        obj = self.model.all_objects.get(pk=pk)
        tenant = getattr(request, 'tenant', None)
        if tenant and hasattr(obj, self.tenant_field) and getattr(obj, self.tenant_field) != tenant:
            messages.error(request, 'Registo não pertence a esta empresa.')
            return redirect(self.redirect_url)
        obj.restore()
        messages.success(request, self.success_message)
        return redirect(self.redirect_url)


class BaseHardDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantFilterMixin, View):
    """View genérica para eliminação permanente de um item da lixeira com tenant."""
    model = None
    redirect_url = None
    protected_error_message = (
        'Nao e possivel eliminar permanentemente este registo.'
    )

    def post(self, request, pk):
        try:
            obj = self.model.all_objects.get(pk=pk)
            tenant = getattr(request, 'tenant', None)
            if tenant and hasattr(obj, self.tenant_field) and getattr(obj, self.tenant_field) != tenant:
                messages.error(request, 'Registo não pertence a esta empresa.')
                return redirect(self.redirect_url)
            obj.hard_delete()
            messages.success(request, self.success_message)
        except ProtectedError:
            messages.error(request, self.protected_error_message)
        return redirect(self.redirect_url)


class SoftDeleteManager(models.Manager):
    """Manager que exclui registos eliminados (soft delete)."""
    def get_queryset(self):
        return super().get_queryset().filter(is_deleted=False)


class SoftDeleteAllManager(models.Manager):
    """Manager que inclui todos os registos (inclusive eliminados)."""
    def get_queryset(self):
        return super().get_queryset()


class SoftDeleteModel(models.Model):
    """Modelo abstracto com soft delete."""
    is_deleted = models.BooleanField(default=False, db_index=True)
    deleted_at = models.DateTimeField(null=True, blank=True)

    objects = SoftDeleteManager()
    all_objects = SoftDeleteAllManager()

    class Meta:
        abstract = True

    def delete(self, using=None, keep_parents=False):
        """Soft delete: marca como eliminado em vez de remover."""
        self.is_deleted = True
        self.deleted_at = timezone.now()
        self.save(update_fields=['is_deleted', 'deleted_at'])

    def hard_delete(self, using=None, keep_parents=False):
        """Remove o registo da base de dados permanentemente."""
        super().delete(using=using, keep_parents=keep_parents)

    def restore(self):
        """Restaura um registo eliminado."""
        self.is_deleted = False
        self.deleted_at = None
        self.save(update_fields=['is_deleted', 'deleted_at'])


class BulkDeleteMixin:
    """Valida permissões antes de eliminação em massa no admin."""

    def delete_queryset(self, request, queryset):
        import logging
        logger = logging.getLogger(__name__)
        if not self.has_delete_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        count = queryset.count()
        logger.info(
            'Bulk delete: user=%s model=%s count=%d',
            request.user, queryset.model._meta.label, count,
        )
        super().delete_queryset(request, queryset)


class SoftDeleteViewMixin:
    """Mixin para DeleteView que faz soft delete em vez de hard delete."""
    def post(self, request, *args, **kwargs):
        try:
            obj = self.get_object()
            obj.delete()
            messages.success(request, self.success_message if hasattr(self, 'success_message') else "Registo excluido com sucesso!")
            return redirect(self.get_success_url())
        except ProtectedError:
            messages.error(request, "Nao e possivel eliminar este registo porque esta a ser utilizado por outros registos.")
            return redirect(self.get_success_url())
