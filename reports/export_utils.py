import io
from django.http import HttpResponse
from django.db.models import Sum
from django.conf import settings
from django.utils import formats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
CELL_FONT = Font(name='Arial', size=10)
CELL_FONT_BOLD = Font(name='Arial', size=10, bold=True)


def apply_header_style(ws, num_columns):
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN


def apply_cell_style(ws, row, num_columns, bold=False):
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT_BOLD if bold else CELL_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical='center')


def auto_width(ws, num_columns):
    for col in range(1, num_columns + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)


def build_excel_response(filename, sheet_title, headers, rows, totals=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    apply_header_style(ws, len(headers))
    for row_num, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        apply_cell_style(ws, row_num, len(headers))
    if totals:
        total_row = ws.max_row + 1
        ws.append(totals)
        apply_cell_style(ws, total_row, len(headers), bold=True)
    auto_width(ws, len(headers))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_pdf_response(filename, title, headers, rows, totals=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    company_style = ParagraphStyle('CompanyHeader', parent=styles['Normal'], fontSize=10, leading=12)
    company_info = settings.COMPANY_INFO
    company_header = [
        Paragraph(f"<b>{company_info['NAME']}</b>", company_style),
        Paragraph(f"{company_info['LOCATION']}", company_style),
        Paragraph(f"Email: {company_info['EMAIL']} | Tel: {company_info['PHONE']}", company_style),
        Spacer(1, 0.5 * cm)
    ]

    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=20)
    elements = company_header + [Paragraph(title, title_style), Spacer(1, 0.5 * cm)]

    available_width = landscape(A4)[0] - 3 * cm
    num_cols = len(headers)
    col_width = available_width / num_cols
    col_widths = [col_width] * num_cols

    table_data = [headers] + rows
    if totals:
        table_data.append(totals)

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    if totals:
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _account_config(account_type):
    if account_type == 'customer':
        return {
            'entity_name_attr': 'customer.name',
            'related_attr': 'outflow',
            'related_label': 'Venda',
            'debit_is_good': False,
            'balance_sign': 1,  # credit - debit
            'sheet_title': 'Extrato Clientes',
            'headers': ['Data', 'Cliente', 'Descricao', 'Debito (Venda)', 'Credito (Pagamento)', 'Saldo Acumulado', 'Situacao'],
        }
    else:
        return {
            'entity_name_attr': 'supplier.name',
            'related_attr': 'inflow',
            'related_label': 'Compra',
            'debit_is_good': True,
            'balance_sign': -1,  # debit - credit
            'sheet_title': 'Extrato Fornecedores',
            'headers': ['Data', 'Fornecedor', 'Descricao', 'Debito (Pago)', 'Credito (Compra)', 'Saldo Acumulado', 'Situacao'],
        }


def build_account_excel(filename, queryset, account_type):
    cfg = _account_config(account_type)
    headers = cfg['headers']
    wb = Workbook()
    ws = wb.active
    ws.title = cfg['sheet_title']
    ws.append(headers)
    apply_header_style(ws, len(headers))

    current_balance = 0
    for row_num, entry in enumerate(queryset, start=2):
        balance_impact = (entry.credit - entry.debit) * cfg['balance_sign']
        current_balance += balance_impact

        related = getattr(entry, cfg['related_attr'], None)
        if related:
            desc = f"{cfg['related_label']} #{related.id} | {related.product.title} (Qtd: {related.quantity})"
            if entry.description:
                desc += f" - {entry.description}"
        else:
            desc = entry.description or ''

        entity_name = entry
        for attr in cfg['entity_name_attr'].split('.'):
            entity_name = getattr(entity_name, attr, '')

        ws.append([
            entry.date.strftime('%d/%m/%Y %H:%M'),
            entity_name,
            desc,
            float(entry.debit) if entry.debit else 0,
            float(entry.credit) if entry.credit else 0,
            float(current_balance),
            'Saldo' if current_balance >= 0 else 'Divida'
        ])

        apply_cell_style(ws, row_num, len(headers))
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).font = Font(name='Arial', size=9)

        if entry.debit > 0:
            color = '008000' if cfg['debit_is_good'] else 'FF0000'
            ws.cell(row=row_num, column=4).font = Font(name='Arial', size=9, color=color)
        if entry.credit > 0:
            color = 'FF0000' if cfg['debit_is_good'] else '008000'
            ws.cell(row=row_num, column=5).font = Font(name='Arial', size=9, color=color)

        balance_color = '008000' if current_balance >= 0 else 'FF0000'
        for col in (6, 7):
            ws.cell(row=row_num, column=col).font = Font(name='Arial', size=9, color=balance_color)
        for col in (4, 5, 6):
            ws.cell(row=row_num, column=col).number_format = '#,##0.00'

    total_row = ws.max_row + 1
    totals = queryset.aggregate(debit=Sum('debit'), credit=Sum('credit'))
    total_debit = totals['debit'] or 0
    total_credit = totals['credit'] or 0
    final_bal = (total_credit - total_debit) * cfg['balance_sign']
    ws.append(['TOTAL', '', '', float(total_debit), float(total_credit), float(final_bal), 'Saldo' if final_bal >= 0 else 'Divida'])
    apply_cell_style(ws, total_row, len(headers), bold=True)
    final_color = '008000' if final_bal >= 0 else 'FF0000'
    for col in (6, 7):
        ws.cell(row=total_row, column=col).font = Font(name='Arial', size=10, bold=True, color=final_color)

    auto_width(ws, len(headers))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_account_pdf(filename, title, queryset, account_type):
    cfg = _account_config(account_type)
    headers = cfg['headers']
    rows = []
    current_balance = 0
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.2, colors.grey),
    ]

    for idx, entry in enumerate(queryset, start=1):
        balance_impact = (entry.credit - entry.debit) * cfg['balance_sign']
        current_balance += balance_impact

        related = getattr(entry, cfg['related_attr'], None)
        if related:
            desc = f"{cfg['related_label']} #{related.id} | {related.product.title} (Qtd: {related.quantity})"
            if entry.description:
                desc += f" - {entry.description}"
        else:
            desc = entry.description or ''

        entity_name = entry
        for attr in cfg['entity_name_attr'].split('.'):
            entity_name = getattr(entity_name, attr, '')

        rows.append([
            entry.date.strftime('%d/%m/%Y'),
            entity_name[:20],
            desc[:50],
            formats.number_format(entry.debit, 2, True) if entry.debit else '0,00',
            formats.number_format(entry.credit, 2, True) if entry.credit else '0,00',
            formats.number_format(current_balance, 2, True),
            'Saldo' if current_balance >= 0 else 'Divida'
        ])

        if entry.debit > 0:
            color = colors.green if cfg['debit_is_good'] else colors.red
            table_styles.append(('TEXTCOLOR', (3, idx), (3, idx), color))
        if entry.credit > 0:
            color = colors.red if cfg['debit_is_good'] else colors.green
            table_styles.append(('TEXTCOLOR', (4, idx), (4, idx), color))

        balance_color = colors.green if current_balance >= 0 else colors.red
        table_styles.append(('TEXTCOLOR', (5, idx), (5, idx), balance_color))
        table_styles.append(('TEXTCOLOR', (6, idx), (6, idx), balance_color))

    totals_agg = queryset.aggregate(debit=Sum('debit'), credit=Sum('credit'))
    total_debit = totals_agg['debit'] or 0
    total_credit = totals_agg['credit'] or 0
    final_balance = (total_credit - total_debit) * cfg['balance_sign']
    totals = ['TOTAL', '', '', formats.number_format(total_debit, 2, True), formats.number_format(total_credit, 2, True), formats.number_format(final_balance, 2, True), 'Saldo' if final_balance >= 0 else 'Divida']

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), margin=1 * cm)
    elements = []
    t = Table([headers] + rows + [totals], colWidths=[2.5 * cm, 3.5 * cm, 9 * cm, 3 * cm, 3 * cm, 3 * cm, 2.5 * cm])
    final_color = colors.green if final_balance >= 0 else colors.red
    t.setStyle(TableStyle(table_styles + [
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (5, -1), (5, -1), final_color),
        ('TEXTCOLOR', (6, -1), (6, -1), final_color),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
