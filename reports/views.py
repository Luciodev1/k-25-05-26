import io
from django.shortcuts import render
from django.http import HttpResponse
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Sum, F, Q
from openpyxl import Workbook
from customers.models import Customer
from suppliers.models import Supplier
from products.models import Product
from outflows.models import Outflow, Delivery
from inflows.models import Inflow
from accounts.models import CustomerAccountEntry, SupplierAccountEntry
from django.utils import formats
from reports.export_utils import (
    build_excel_response, build_pdf_response,
    build_account_excel, build_account_pdf,
    apply_header_style, apply_cell_style, auto_width,
)

ASYNC_EXPORT_THRESHOLD = 1000


def _get_user_email(request):
    return getattr(request.user, 'email', None)


def _dispatch_async_if_large(queryset, request, async_task, export_format, filters, task_kwargs=None):
    count = queryset.count()
    if count <= ASYNC_EXPORT_THRESHOLD:
        return None
    tenant_id = str(getattr(request, 'tenant', None) or '')
    task_kwargs = task_kwargs or {}
    task = async_task.delay(
        _get_user_email(request),
        tenant_id,
        filters,
        export_format,
        **task_kwargs,
    )
    return render(request, 'report_processing.html', {
        'task_id': task.id,
        'export_format': export_format.upper(),
        'record_count': count,
    })


def _get_filters(request):
    def clean_value(val):
        if val in (None, '', 'None'):
            return None
        return val

    return {
        'start_date': clean_value(request.GET.get('start_date')),
        'end_date': clean_value(request.GET.get('end_date')),
        'customer_id': clean_value(request.GET.get('customer')),
        'supplier_id': clean_value(request.GET.get('supplier')),
        'product_id': clean_value(request.GET.get('product')),
        'status': clean_value(request.GET.get('status')),
    }


from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import PermissionDenied
from django_ratelimit.decorators import ratelimit


@login_required
def report_index(request):
    if not (request.user.has_perm('outflows.view_outflow') or 
            request.user.has_perm('accounts.view_customeraccountentry') or 
            request.user.has_perm('accounts.view_supplieraccountentry')):
        raise PermissionDenied
    return render(request, 'report_index.html')


@login_required
@permission_required('outflows.view_outflow', raise_exception=True)
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def outflows_by_customer_report(request):
    filters = _get_filters(request)
    tenant = getattr(request, 'tenant', None)
    base_qs = Outflow.objects.select_related('product', 'customer')
    if tenant:
        base_qs = base_qs.filter(tenant=tenant)
    queryset = base_qs.all()

    if filters['start_date']:
        queryset = queryset.filter(created_at__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(created_at__lte=filters['end_date'])
    if filters['customer_id']:
        queryset = queryset.filter(customer_id=filters['customer_id'])
    if filters['product_id']:
        queryset = queryset.filter(product_id=filters['product_id'])

    export = request.GET.get('export')

    if export in ('excel', 'pdf'):
        from reports.export_tasks import async_outflows_by_customer_report
        async_response = _dispatch_async_if_large(
            queryset, request, async_outflows_by_customer_report,
            export, filters,
        )
        if async_response:
            return async_response

    if export == 'excel':
        headers = ['Data', 'Cliente', 'Produto', 'Quantidade', 'Qtd Entregue', 'Qtd Pendente', 'Estado']
        rows = [[o.created_at.strftime('%d/%m/%Y'), o.customer.name, o.product.title,
                 o.quantity, o.quantity_delivered, o.quantity_pending, o.status_display] for o in queryset]
        totals_agg = queryset.aggregate(qty=Sum('quantity'), delivered=Sum('quantity_delivered'))
        totals = ['TOTAL', '', '', totals_agg['qty'] or 0, totals_agg['delivered'] or 0, '', '']
        return build_excel_response('saidas_por_cliente.xlsx', 'Saidas por Cliente', headers, rows, totals)

    if export == 'pdf':
        headers = ['Data', 'Cliente', 'Produto', 'Qtd', 'Entregue', 'Pendente', 'Estado']
        rows = []
        for o in queryset:
            rows.append([
                o.created_at.strftime('%d/%m/%Y'),
                o.customer.name,
                o.product.title,
                str(o.quantity),
                str(o.quantity_delivered),
                str(o.quantity_pending),
                o.status_display,
            ])
        totals_agg = queryset.aggregate(
            qty=Sum('quantity'),
            delivered=Sum('quantity_delivered')
        )
        totals = ['TOTAL', '', '', str(totals_agg['qty'] or 0), str(totals_agg['delivered'] or 0), '', '']
        return build_pdf_response('saidas_por_cliente.pdf', 'Saidas por Cliente', headers, rows, totals)

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'outflows': page_obj,
        'page_obj': page_obj,
        'customers': Customer.objects.filter(tenant=tenant).only('id', 'name') if tenant else Customer.objects.only('id', 'name'),
        'products': Product.objects.filter(tenant=tenant).only('id', 'title') if tenant else Product.objects.only('id', 'title'),
        'filters': filters,
    }
    return render(request, 'report_outflows_by_customer.html', context)


@login_required
@permission_required('outflows.view_outflow', raise_exception=True)
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def deliveries_report(request):
    filters = _get_filters(request)
    tenant = getattr(request, 'tenant', None)
    base_qs = Delivery.objects.select_related('outflow__product', 'outflow__customer')
    if tenant:
        base_qs = base_qs.filter(tenant=tenant)
    queryset = base_qs.all()

    if filters['start_date']:
        queryset = queryset.filter(delivered_at__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(delivered_at__lte=filters['end_date'])
    if filters['customer_id']:
        queryset = queryset.filter(outflow__customer_id=filters['customer_id'])
    if filters['product_id']:
        queryset = queryset.filter(outflow__product_id=filters['product_id'])
    if filters['status']:
        if filters['status'] == 'pending':
            queryset = queryset.filter(outflow__quantity_delivered__lt=F('outflow__quantity'))
        elif filters['status'] == 'delivered':
            queryset = queryset.filter(outflow__quantity_delivered=F('outflow__quantity'))

    export = request.GET.get('export')

    if export in ('excel', 'pdf'):
        from reports.export_tasks import async_deliveries_report
        async_response = _dispatch_async_if_large(
            queryset, request, async_deliveries_report,
            export, filters,
        )
        if async_response:
            return async_response

    if export == 'excel':
        headers = ['Data Entrega', 'Cliente', 'Produto', 'Qtd Entregue', 'Descrição']
        rows = [[d.delivered_at.strftime('%d/%m/%Y'), d.outflow.customer.name,
                 d.outflow.product.title, d.quantity, d.description or ''] for d in queryset]
        total_qty = queryset.aggregate(total=Sum('quantity'))['total'] or 0
        totals = ['TOTAL', '', '', total_qty, '']
        return build_excel_response('entregas.xlsx', 'Entregas', headers, rows, totals)

    if export == 'pdf':
        headers = ['Data', 'Cliente', 'Produto', 'Qtd', 'Descricao']
        rows = []
        for d in queryset:
            rows.append([
                d.delivered_at.strftime('%d/%m/%Y'),
                d.outflow.customer.name,
                d.outflow.product.title,
                str(d.quantity),
                d.description or '',
            ])
        total_qty = queryset.aggregate(total=Sum('quantity'))['total'] or 0
        totals = ['TOTAL', '', '', str(total_qty), '']
        return build_pdf_response('entregas.pdf', 'Relatorio de Entregas', headers, rows, totals)

    outflows_pending = Outflow.objects.filter(quantity_delivered__lt=F('quantity')).select_related('customer', 'product')
    if tenant:
        outflows_pending = outflows_pending.filter(tenant=tenant)
    if filters['customer_id']:
        outflows_pending = outflows_pending.filter(customer_id=filters['customer_id'])
    if filters['product_id']:
        outflows_pending = outflows_pending.filter(product_id=filters['product_id'])

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'deliveries': page_obj,
        'page_obj': page_obj,
        'outflows_pending': outflows_pending,
        'customers': Customer.objects.filter(tenant=tenant).only('id', 'name') if tenant else Customer.objects.only('id', 'name'),
        'products': Product.objects.filter(tenant=tenant).only('id', 'title') if tenant else Product.objects.only('id', 'title'),
        'filters': filters,
    }
    return render(request, 'report_deliveries.html', context)


@login_required
@permission_required('accounts.view_customeraccountentry', raise_exception=True)
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def customer_account_report(request):
    filters = _get_filters(request)
    tenant = getattr(request, 'tenant', None)
    base_qs = CustomerAccountEntry.objects.select_related('customer', 'outflow__product')
    if tenant:
        base_qs = base_qs.filter(tenant=tenant)
    queryset = base_qs.all()

    if filters['start_date']:
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(date__lte=filters['end_date'])
    if filters['customer_id']:
        queryset = queryset.filter(customer_id=filters['customer_id'])

    export = request.GET.get('export')

    if export in ('excel', 'pdf'):
        from reports.export_tasks import async_customer_account_report
        async_response = _dispatch_async_if_large(
            queryset, request, async_customer_account_report,
            export, filters,
        )
        if async_response:
            return async_response

    if export == 'excel':
        return build_account_excel('extrato_clientes.xlsx', queryset, 'customer')

    if export == 'pdf':
        return build_account_pdf('extrato_clientes.pdf', 'Extrato de Clientes', queryset, 'customer')

    annotated_qs = queryset.annotate(entry_balance=F('credit') - F('debit'))
    paginator = Paginator(annotated_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'entries': page_obj,
        'page_obj': page_obj,
        'customers': Customer.objects.filter(tenant=tenant).only('id', 'name') if tenant else Customer.objects.only('id', 'name'),
        'filters': filters,
        'total_debit': queryset.aggregate(total=Sum('debit'))['total'] or 0,
        'total_credit': queryset.aggregate(total=Sum('credit'))['total'] or 0,
    }
    context['balance'] = context['total_credit'] - context['total_debit']
    
    if request.headers.get('HX-Request'):
        return render(request, 'report_customer_account_table.html', context)
    
    return render(request, 'report_customer_account.html', context)


@login_required
@permission_required('accounts.view_supplieraccountentry', raise_exception=True)
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def supplier_account_report(request):
    filters = _get_filters(request)
    tenant = getattr(request, 'tenant', None)
    base_qs = SupplierAccountEntry.objects.select_related('supplier', 'inflow__product')
    if tenant:
        base_qs = base_qs.filter(tenant=tenant)
    queryset = base_qs.all()

    if filters['start_date']:
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters['end_date']:
        queryset = queryset.filter(date__lte=filters['end_date'])
    if filters['supplier_id']:
        queryset = queryset.filter(supplier_id=filters['supplier_id'])

    export = request.GET.get('export')

    if export in ('excel', 'pdf'):
        from reports.export_tasks import async_supplier_account_report
        async_response = _dispatch_async_if_large(
            queryset, request, async_supplier_account_report,
            export, filters,
        )
        if async_response:
            return async_response

    if export == 'excel':
        return build_account_excel('extrato_fornecedores.xlsx', queryset, 'supplier')

    if export == 'pdf':
        return build_account_pdf('extrato_fornecedores.pdf', 'Extrato de Fornecedores', queryset, 'supplier')

    annotated_qs = queryset.annotate(entry_balance=F('debit') - F('credit'))
    paginator = Paginator(annotated_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'entries': page_obj,
        'page_obj': page_obj,
        'suppliers': Supplier.objects.filter(tenant=tenant).only('id', 'name') if tenant else Supplier.objects.only('id', 'name'),
        'filters': filters,
        'total_debit': queryset.aggregate(total=Sum('debit'))['total'] or 0,
        'total_credit': queryset.aggregate(total=Sum('credit'))['total'] or 0,
    }
    context['balance'] = context['total_debit'] - context['total_credit']
    return render(request, 'report_supplier_account.html', context)


@login_required
@ratelimit(key='user_or_ip', rate='10/m', method='GET', block=True)
def balances_report(request):
    if not (request.user.has_perm('accounts.view_customeraccountentry') or 
            request.user.has_perm('accounts.view_supplieraccountentry')):
        raise PermissionDenied
    filters = _get_filters(request)
    tenant = getattr(request, 'tenant', None)
    
    customer_filter = Q()
    supplier_filter = Q()
    if filters['start_date']:
        customer_filter &= Q(account_entries__date__gte=filters['start_date'])
        supplier_filter &= Q(account_entries__date__gte=filters['start_date'])
    if filters['end_date']:
        customer_filter &= Q(account_entries__date__lte=filters['end_date'])
        supplier_filter &= Q(account_entries__date__lte=filters['end_date'])

    c_base = Customer.objects
    s_base = Supplier.objects
    if tenant:
        c_base = c_base.filter(tenant=tenant)
        s_base = s_base.filter(tenant=tenant)

    customer_balances = c_base.annotate(
        total_debit=Sum('account_entries__debit', filter=customer_filter, default=0),
        total_credit=Sum('account_entries__credit', filter=customer_filter, default=0),
        balance=F('total_credit') - F('total_debit'),
    ).order_by('name')

    supplier_balances = s_base.annotate(
        total_debit=Sum('account_entries__debit', filter=supplier_filter, default=0),
        total_credit=Sum('account_entries__credit', filter=supplier_filter, default=0),
        balance=F('total_debit') - F('total_credit'),
    ).order_by('name')

    export = request.GET.get('export')
    section = request.GET.get('section', 'all')

    if export in ('excel', 'pdf'):
        from reports.export_tasks import async_balances_report
        count = (customer_balances.count() if section in ('all', 'customers') else 0) + \
                (supplier_balances.count() if section in ('all', 'suppliers') else 0)
        if count > ASYNC_EXPORT_THRESHOLD:
            task = async_balances_report.delay(
                _get_user_email(request),
                str(getattr(request, 'tenant', None) or ''),
                filters,
                export,
                section,
            )
            return render(request, 'report_processing.html', {
                'task_id': task.id,
                'export_format': export.upper(),
                'record_count': count,
            })

    if export == 'excel':
        wb = Workbook()

        if section in ('all', 'customers'):
            ws = wb.active
            ws.title = 'Saldos Clientes'
            headers = ['Cliente', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            ws.append(headers)
            apply_header_style(ws, len(headers))
            for row_num, c in enumerate(customer_balances, start=2):
                ws.append([c.name, float(c.total_debit), float(c.total_credit), float(c.balance), 'Saldo' if c.balance >= 0 else 'Dívida'])
                ws.cell(row=row_num, column=2).number_format = '#,##0.00'
                ws.cell(row=row_num, column=3).number_format = '#,##0.00'
                ws.cell(row=row_num, column=4).number_format = '#,##0.00'
                apply_cell_style(ws, row_num, len(headers))
            total_row = ws.max_row + 1
            total_d = customer_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = customer_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = customer_balances.aggregate(total=Sum('balance'))['total'] or 0
            ws.append([
                'TOTAL',
                float(total_d),
                float(total_c),
                float(total_b),
                ''
            ])
            ws.cell(row=total_row, column=2).number_format = '#,##0.00'
            ws.cell(row=total_row, column=3).number_format = '#,##0.00'
            ws.cell(row=total_row, column=4).number_format = '#,##0.00'
            apply_cell_style(ws, total_row, len(headers), bold=True)
            auto_width(ws, len(headers))

        if section in ('all', 'suppliers'):
            if section == 'all':
                ws2 = wb.create_sheet('Saldos Fornecedores')
            else:
                ws2 = wb.active
                ws2.title = 'Saldos Fornecedores'
            headers = ['Fornecedor', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            ws2.append(headers)
            apply_header_style(ws2, len(headers))
            for row_num, s in enumerate(supplier_balances, start=2):
                ws2.append([s.name, float(s.total_debit), float(s.total_credit), float(s.balance), 'Saldo' if s.balance >= 0 else 'Dívida'])
                ws2.cell(row=row_num, column=2).number_format = '#,##0.00'
                ws2.cell(row=row_num, column=3).number_format = '#,##0.00'
                ws2.cell(row=row_num, column=4).number_format = '#,##0.00'
                apply_cell_style(ws2, row_num, len(headers))
            total_row = ws2.max_row + 1
            total_d = supplier_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = supplier_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = supplier_balances.aggregate(total=Sum('balance'))['total'] or 0
            ws2.append([
                'TOTAL',
                float(total_d),
                float(total_c),
                float(total_b),
                ''
            ])
            ws2.cell(row=total_row, column=2).number_format = '#,##0.00'
            ws2.cell(row=total_row, column=3).number_format = '#,##0.00'
            ws2.cell(row=total_row, column=4).number_format = '#,##0.00'
            apply_cell_style(ws2, total_row, len(headers), bold=True)
            auto_width(ws2, len(headers))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="saldos.xlsx"'
        return response

    if export == 'pdf':
        if section == 'suppliers':
            headers = ['Fornecedor', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            rows = []
            for s in supplier_balances:
                rows.append([
                    s.name,
                    formats.number_format(s.total_debit, decimal_pos=2, use_l10n=True),
                    formats.number_format(s.total_credit, decimal_pos=2, use_l10n=True),
                    formats.number_format(s.balance, decimal_pos=2, use_l10n=True),
                    'Saldo' if s.balance >= 0 else 'Dívida'
                ])
            total_d = supplier_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = supplier_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = supplier_balances.aggregate(total=Sum('balance'))['total'] or 0
            totals = [
                'TOTAL',
                formats.number_format(total_d, decimal_pos=2, use_l10n=True),
                formats.number_format(total_c, decimal_pos=2, use_l10n=True),
                formats.number_format(total_b, decimal_pos=2, use_l10n=True), ''
            ]
            return build_pdf_response('saldos_fornecedores.pdf', 'Saldos de Fornecedores', headers, rows, totals)
        else:
            headers = ['Cliente', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            rows = []
            for c in customer_balances:
                rows.append([
                    c.name,
                    formats.number_format(c.total_debit, decimal_pos=2, use_l10n=True),
                    formats.number_format(c.total_credit, decimal_pos=2, use_l10n=True),
                    formats.number_format(c.balance, decimal_pos=2, use_l10n=True),
                    'Saldo' if c.balance >= 0 else 'Dívida'
                ])
            total_d = customer_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = customer_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = customer_balances.aggregate(total=Sum('balance'))['total'] or 0
            totals = [
                'TOTAL',
                formats.number_format(total_d, decimal_pos=2, use_l10n=True),
                formats.number_format(total_c, decimal_pos=2, use_l10n=True),
                formats.number_format(total_b, decimal_pos=2, use_l10n=True), ''
            ]
            return build_pdf_response('saldos_clientes.pdf', 'Saldos de Clientes', headers, rows, totals)

    context = {
        'customer_balances': customer_balances,
        'supplier_balances': supplier_balances,
        'filters': filters,
    }
    return render(request, 'report_balances.html', context)


@login_required
def task_status(request, task_id):
    from tenants.models import TenantUser
    tenant = getattr(request, 'tenant', None)
    if tenant:
        if not TenantUser.objects.filter(user=request.user, tenant=tenant).exists():
            raise PermissionDenied
    if not (request.user.has_perm('outflows.view_outflow') or
            request.user.has_perm('accounts.view_customeraccountentry') or
            request.user.has_perm('accounts.view_supplieraccountentry')):
        raise PermissionDenied
    from celery.result import AsyncResult
    from django.http import JsonResponse
    result = AsyncResult(task_id)
    ready = result.ready()
    payload = {'task_id': task_id, 'status': result.status}
    download_url = None
    if ready and result.result and isinstance(result.result, dict) and result.result.get('status') == 'ok':
        path = result.result.get('path', '')
        if path:
            download_url = reverse('reports:report_download', kwargs={'task_id': task_id})
            payload['result'] = result.result
            payload['download_url'] = download_url
    if request.headers.get('HX-Request'):
        if ready:
            if download_url:
                return render(request, 'report_download_ready.html', {
                    'download_url': download_url,
                    'task_id': task_id,
                })
            return HttpResponse('<span class="text-danger">Falha na exportação.</span>')
        return HttpResponse(
            '<div hx-get="{}" hx-trigger="load delay:2s" hx-swap="outerHTML">'
            '<span class="text-muted">A processar...</span></div>'.format(
                reverse('reports:report_task_status', kwargs={'task_id': task_id})
            )
        )
    payload['download_url'] = download_url
    return JsonResponse(payload)


@login_required
def report_download(request, task_id):
    from django.http import FileResponse, Http404
    from django.core.files.storage import default_storage
    from celery.result import AsyncResult
    result = AsyncResult(task_id)
    if not result.ready() or not result.result:
        raise Http404('Exportação não encontrada ou ainda em processamento.')
    path = result.result.get('path', '') if isinstance(result.result, dict) else ''
    if not path or not default_storage.exists(path):
        raise Http404('Ficheiro não encontrado ou expirado.')
    from pathlib import Path
    filename = Path(path).name
    response = FileResponse(default_storage.open(path, 'rb'), as_attachment=True, filename=filename)
    return response
