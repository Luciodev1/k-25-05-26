import io
import logging
from typing import Any

from celery import shared_task
from celery.app.task import Task
from django.conf import settings
from django.core.files.storage import default_storage
from django.db.models import Sum, F, Q
from django.utils import formats
from openpyxl import Workbook

from reports.export_utils import apply_header_style, apply_cell_style, auto_width

logger = logging.getLogger(__name__)


def _notify(user_email: str | None, task_name: str, path: str | None = None) -> None:
    if not user_email:
        return
    from app.tasks import notify_task_completion
    notify_task_completion.delay(user_email, task_name, path)


@shared_task(bind=True)
def async_outflows_by_customer_report(
    self: Task,
    user_email: str | None,
    tenant_id: str | None,
    filters: dict[str, Any],
    export_format: str,
) -> dict[str, Any]:
    from outflows.models import Outflow

    base_qs = Outflow.objects.select_related('product', 'customer')
    if tenant_id:
        base_qs = base_qs.filter(tenant_id=tenant_id)
    queryset = base_qs.all()

    if filters.get('start_date'):
        queryset = queryset.filter(created_at__gte=filters['start_date'])
    if filters.get('end_date'):
        queryset = queryset.filter(created_at__lte=filters['end_date'])
    if filters.get('customer_id'):
        queryset = queryset.filter(customer_id=filters['customer_id'])
    if filters.get('product_id'):
        queryset = queryset.filter(product_id=filters['product_id'])

    if export_format == 'excel':
        headers = ['Data', 'Cliente', 'Produto', 'Quantidade', 'Qtd Entregue', 'Qtd Pendente', 'Estado']
        rows = [[o.created_at.strftime('%d/%m/%Y'), o.customer.name, o.product.title,
                 o.quantity, o.quantity_delivered, o.quantity_pending, o.status_display] for o in queryset]
        totals_agg = queryset.aggregate(qty=Sum('quantity'), delivered=Sum('quantity_delivered'))
        totals = ['TOTAL', '', '', totals_agg['qty'] or 0, totals_agg['delivered'] or 0, '', '']

        wb = Workbook()
        ws = wb.active
        ws.title = 'Saidas por Cliente'
        ws.append(headers)
        apply_header_style(ws, len(headers))
        for row_num, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            apply_cell_style(ws, row_num, len(headers))
        total_row = ws.max_row + 1
        ws.append(totals)
        apply_cell_style(ws, total_row, len(headers), bold=True)
        auto_width(ws, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        path = f'exports/{self.request.id}_saidas_por_cliente.xlsx'
        buffer.seek(0)
        default_storage.save(path, buffer)
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        headers = ['Data', 'Cliente', 'Produto', 'Qtd', 'Entregue', 'Pendente', 'Estado']
        rows = []
        for o in queryset:
            rows.append([
                o.created_at.strftime('%d/%m/%Y'),
                o.customer.name,
                o.product.title,
                str(o.quantity),
                str(o.quantity_delivered),
                str(o.quantity_pending),
                o.status_display,
            ])
        totals_agg = queryset.aggregate(qty=Sum('quantity'), delivered=Sum('quantity_delivered'))
        totals = ['TOTAL', '', '', str(totals_agg['qty'] or 0), str(totals_agg['delivered'] or 0), '', '']

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        company_info = settings.COMPANY_INFO
        company_style = ParagraphStyle('CompanyHeader', parent=styles['Normal'], fontSize=10, leading=12)
        elements = [
            Paragraph(f"<b>{company_info['NAME']}</b>", company_style),
            Paragraph(f"{company_info['LOCATION']}", company_style),
            Paragraph(f"Email: {company_info['EMAIL']} | Tel: {company_info['PHONE']}", company_style),
            Spacer(1, 0.5 * cm),
            Paragraph('Saidas por Cliente', styles['Title']),
            Spacer(1, 0.5 * cm),
        ]
        available_width = landscape(A4)[0] - 3 * cm
        col_width = available_width / len(headers)
        table_data = [headers] + rows + [totals]
        table = Table(table_data, colWidths=[col_width] * len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)

        path = f'exports/{self.request.id}_saidas_por_cliente.pdf'
        buffer.seek(0)
        default_storage.save(path, buffer)

    logger.info('Export async concluído: %s (%d registos)', path, queryset.count())
    _notify(user_email, f'Export Saidas por Cliente ({export_format.upper()})', path)
    return {'status': 'ok', 'path': path}


@shared_task(bind=True)
def async_deliveries_report(
    self: Task,
    user_email: str | None,
    tenant_id: str | None,
    filters: dict[str, Any],
    export_format: str,
) -> dict[str, Any]:
    from outflows.models import Delivery

    base_qs = Delivery.objects.select_related('outflow__product', 'outflow__customer')
    if tenant_id:
        base_qs = base_qs.filter(tenant=tenant_id)
    queryset = base_qs.all()

    if filters.get('start_date'):
        queryset = queryset.filter(delivered_at__gte=filters['start_date'])
    if filters.get('end_date'):
        queryset = queryset.filter(delivered_at__lte=filters['end_date'])
    if filters.get('customer_id'):
        queryset = queryset.filter(outflow__customer_id=filters['customer_id'])
    if filters.get('product_id'):
        queryset = queryset.filter(outflow__product_id=filters['product_id'])
    if filters.get('status'):
        if filters['status'] == 'pending':
            queryset = queryset.filter(outflow__quantity_delivered__lt=F('outflow__quantity'))
        elif filters['status'] == 'delivered':
            queryset = queryset.filter(outflow__quantity_delivered=F('outflow__quantity'))

    if export_format == 'excel':
        headers = ['Data Entrega', 'Cliente', 'Produto', 'Qtd Entregue', 'Descrição']
        rows = [[d.delivered_at.strftime('%d/%m/%Y'), d.outflow.customer.name,
                 d.outflow.product.title, d.quantity, d.description or ''] for d in queryset]
        total_qty = queryset.aggregate(total=Sum('quantity'))['total'] or 0
        totals = ['TOTAL', '', '', total_qty, '']

        wb = Workbook()
        ws = wb.active
        ws.title = 'Entregas'
        ws.append(headers)
        apply_header_style(ws, len(headers))
        for row_num, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            apply_cell_style(ws, row_num, len(headers))
        total_row = ws.max_row + 1
        ws.append(totals)
        apply_cell_style(ws, total_row, len(headers), bold=True)
        auto_width(ws, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        path = f'exports/{self.request.id}_entregas.xlsx'
        buffer.seek(0)
        default_storage.save(path, buffer)
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        headers = ['Data', 'Cliente', 'Produto', 'Qtd', 'Descricao']
        rows = []
        for d in queryset:
            rows.append([
                d.delivered_at.strftime('%d/%m/%Y'),
                d.outflow.customer.name,
                d.outflow.product.title,
                str(d.quantity),
                d.description or '',
            ])
        total_qty = queryset.aggregate(total=Sum('quantity'))['total'] or 0
        totals = ['TOTAL', '', '', str(total_qty), '']

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        company_info = settings.COMPANY_INFO
        company_style = ParagraphStyle('CompanyHeader', parent=styles['Normal'], fontSize=10, leading=12)
        elements = [
            Paragraph(f"<b>{company_info['NAME']}</b>", company_style),
            Paragraph(f"{company_info['LOCATION']}", company_style),
            Paragraph(f"Email: {company_info['EMAIL']} | Tel: {company_info['PHONE']}", company_style),
            Spacer(1, 0.5 * cm),
            Paragraph('Relatorio de Entregas', styles['Title']),
            Spacer(1, 0.5 * cm),
        ]
        available_width = landscape(A4)[0] - 3 * cm
        col_width = available_width / len(headers)
        table_data = [headers] + rows + [totals]
        table = Table(table_data, colWidths=[col_width] * len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        path = f'exports/{self.request.id}_entregas.pdf'
        buffer.seek(0)
        default_storage.save(path, buffer)

    logger.info('Export async concluído: %s', path)
    _notify(user_email, f'Export Entregas ({export_format.upper()})', path)
    return {'status': 'ok', 'path': path}


@shared_task(bind=True)
def async_customer_account_report(
    self: Task,
    user_email: str | None,
    tenant_id: str | None,
    filters: dict[str, Any],
    export_format: str,
) -> dict[str, Any]:
    from accounts.models import CustomerAccountEntry
    from reports.export_utils import build_account_excel, build_account_pdf

    base_qs = CustomerAccountEntry.objects.select_related('customer', 'outflow__product')
    if tenant_id:
        base_qs = base_qs.filter(tenant=tenant_id)
    queryset = base_qs.all()

    if filters.get('start_date'):
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters.get('end_date'):
        queryset = queryset.filter(date__lte=filters['end_date'])
    if filters.get('customer_id'):
        queryset = queryset.filter(customer_id=filters['customer_id'])

    if export_format == 'excel':
        response = build_account_excel('extrato_clientes.xlsx', queryset, 'customer')
        buffer = io.BytesIO(response.content)
        path = f'exports/{self.request.id}_extrato_clientes.xlsx'
        buffer.seek(0)
        default_storage.save(path, buffer)
    else:
        response = build_account_pdf('extrato_clientes.pdf', 'Extrato de Clientes', queryset, 'customer')
        buffer = io.BytesIO(response.content)
        path = f'exports/{self.request.id}_extrato_clientes.pdf'
        buffer.seek(0)
        default_storage.save(path, buffer)

    logger.info('Export async concluído: %s', path)
    _notify(user_email, f'Export Extrato Clientes ({export_format.upper()})', path)
    return {'status': 'ok', 'path': path}


@shared_task(bind=True)
def async_supplier_account_report(
    self: Task,
    user_email: str | None,
    tenant_id: str | None,
    filters: dict[str, Any],
    export_format: str,
) -> dict[str, Any]:
    from accounts.models import SupplierAccountEntry
    from reports.export_utils import build_account_excel, build_account_pdf

    base_qs = SupplierAccountEntry.objects.select_related('supplier', 'inflow__product')
    if tenant_id:
        base_qs = base_qs.filter(tenant=tenant_id)
    queryset = base_qs.all()

    if filters.get('start_date'):
        queryset = queryset.filter(date__gte=filters['start_date'])
    if filters.get('end_date'):
        queryset = queryset.filter(date__lte=filters['end_date'])
    if filters.get('supplier_id'):
        queryset = queryset.filter(supplier_id=filters['supplier_id'])

    if export_format == 'excel':
        response = build_account_excel('extrato_fornecedores.xlsx', queryset, 'supplier')
        buffer = io.BytesIO(response.content)
        path = f'exports/{self.request.id}_extrato_fornecedores.xlsx'
        buffer.seek(0)
        default_storage.save(path, buffer)
    else:
        response = build_account_pdf('extrato_fornecedores.pdf', 'Extrato de Fornecedores', queryset, 'supplier')
        buffer = io.BytesIO(response.content)
        path = f'exports/{self.request.id}_extrato_fornecedores.pdf'
        buffer.seek(0)
        default_storage.save(path, buffer)

    logger.info('Export async concluído: %s', path)
    _notify(user_email, f'Export Extrato Fornecedores ({export_format.upper()})', path)
    return {'status': 'ok', 'path': path}


@shared_task(bind=True)
def async_balances_report(
    self: Task,
    user_email: str | None,
    tenant_id: str | None,
    filters: dict[str, Any],
    export_format: str,
    section: str = 'all',
) -> dict[str, Any]:
    from customers.models import Customer
    from suppliers.models import Supplier
    from django.http import HttpResponse

    customer_filter = Q()
    supplier_filter = Q()
    if filters.get('start_date'):
        customer_filter &= Q(account_entries__date__gte=filters['start_date'])
        supplier_filter &= Q(account_entries__date__gte=filters['start_date'])
    if filters.get('end_date'):
        customer_filter &= Q(account_entries__date__lte=filters['end_date'])
        supplier_filter &= Q(account_entries__date__lte=filters['end_date'])

    c_base = Customer.objects
    s_base = Supplier.objects
    if tenant_id:
        c_base = c_base.filter(tenant=tenant_id)
        s_base = s_base.filter(tenant=tenant_id)

    customer_balances = c_base.annotate(
        total_debit=Sum('account_entries__debit', filter=customer_filter, default=0),
        total_credit=Sum('account_entries__credit', filter=customer_filter, default=0),
        balance=F('total_credit') - F('total_debit'),
    ).order_by('name')

    supplier_balances = s_base.annotate(
        total_debit=Sum('account_entries__debit', filter=supplier_filter, default=0),
        total_credit=Sum('account_entries__credit', filter=supplier_filter, default=0),
        balance=F('total_debit') - F('total_credit'),
    ).order_by('name')

    if export_format == 'excel':
        wb = Workbook()
        if section in ('all', 'customers'):
            ws = wb.active
            ws.title = 'Saldos Clientes'
            headers = ['Cliente', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            ws.append(headers)
            apply_header_style(ws, len(headers))
            for row_num, c in enumerate(customer_balances, start=2):
                ws.append([c.name, float(c.total_debit), float(c.total_credit), float(c.balance), 'Saldo' if c.balance >= 0 else 'Dívida'])
                ws.cell(row=row_num, column=2).number_format = '#,##0.00'
                ws.cell(row=row_num, column=3).number_format = '#,##0.00'
                ws.cell(row=row_num, column=4).number_format = '#,##0.00'
                apply_cell_style(ws, row_num, len(headers))
            total_row = ws.max_row + 1
            total_d = customer_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = customer_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = customer_balances.aggregate(total=Sum('balance'))['total'] or 0
            ws.append(['TOTAL', float(total_d), float(total_c), float(total_b), ''])
            ws.cell(row=total_row, column=2).number_format = '#,##0.00'
            ws.cell(row=total_row, column=3).number_format = '#,##0.00'
            ws.cell(row=total_row, column=4).number_format = '#,##0.00'
            apply_cell_style(ws, total_row, len(headers), bold=True)
            auto_width(ws, len(headers))

        if section in ('all', 'suppliers'):
            ws2 = wb.create_sheet('Saldos Fornecedores') if section == 'all' else wb.active
            ws2.title = 'Saldos Fornecedores'
            headers = ['Fornecedor', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            ws2.append(headers)
            apply_header_style(ws2, len(headers))
            for row_num, s in enumerate(supplier_balances, start=2):
                ws2.append([s.name, float(s.total_debit), float(s.total_credit), float(s.balance), 'Saldo' if s.balance >= 0 else 'Dívida'])
                ws2.cell(row=row_num, column=2).number_format = '#,##0.00'
                ws2.cell(row=row_num, column=3).number_format = '#,##0.00'
                ws2.cell(row=row_num, column=4).number_format = '#,##0.00'
                apply_cell_style(ws2, row_num, len(headers))
            total_row = ws2.max_row + 1
            total_d = supplier_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = supplier_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = supplier_balances.aggregate(total=Sum('balance'))['total'] or 0
            ws2.append(['TOTAL', float(total_d), float(total_c), float(total_b), ''])
            ws2.cell(row=total_row, column=2).number_format = '#,##0.00'
            ws2.cell(row=total_row, column=3).number_format = '#,##0.00'
            ws2.cell(row=total_row, column=4).number_format = '#,##0.00'
            apply_cell_style(ws2, total_row, len(headers), bold=True)
            auto_width(ws2, len(headers))

        buffer = io.BytesIO()
        wb.save(buffer)
        path = f'exports/{self.request.id}_saldos.xlsx'
        buffer.seek(0)
        default_storage.save(path, buffer)
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        if section == 'suppliers':
            headers = ['Fornecedor', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            rows = []
            for s in supplier_balances:
                rows.append([
                    s.name,
                    formats.number_format(s.total_debit, decimal_pos=2, use_l10n=True),
                    formats.number_format(s.total_credit, decimal_pos=2, use_l10n=True),
                    formats.number_format(s.balance, decimal_pos=2, use_l10n=True),
                    'Saldo' if s.balance >= 0 else 'Dívida'
                ])
            total_d = supplier_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = supplier_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = supplier_balances.aggregate(total=Sum('balance'))['total'] or 0
            totals = ['TOTAL', formats.number_format(total_d, decimal_pos=2, use_l10n=True),
                      formats.number_format(total_c, decimal_pos=2, use_l10n=True),
                      formats.number_format(total_b, decimal_pos=2, use_l10n=True), '']
            filename = 'saldos_fornecedores.pdf'
        else:
            headers = ['Cliente', 'Total Débito', 'Total Crédito', 'Saldo Final', 'Situação']
            rows = []
            for c in customer_balances:
                rows.append([
                    c.name,
                    formats.number_format(c.total_debit, decimal_pos=2, use_l10n=True),
                    formats.number_format(c.total_credit, decimal_pos=2, use_l10n=True),
                    formats.number_format(c.balance, decimal_pos=2, use_l10n=True),
                    'Saldo' if c.balance >= 0 else 'Dívida'
                ])
            total_d = customer_balances.aggregate(total=Sum('total_debit'))['total'] or 0
            total_c = customer_balances.aggregate(total=Sum('total_credit'))['total'] or 0
            total_b = customer_balances.aggregate(total=Sum('balance'))['total'] or 0
            totals = ['TOTAL', formats.number_format(total_d, decimal_pos=2, use_l10n=True),
                      formats.number_format(total_c, decimal_pos=2, use_l10n=True),
                      formats.number_format(total_b, decimal_pos=2, use_l10n=True), '']
            filename = 'saldos_clientes.pdf'

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        company_info = settings.COMPANY_INFO
        company_style = ParagraphStyle('CompanyHeader', parent=styles['Normal'], fontSize=10, leading=12)
        elements = [
            Paragraph(f"<b>{company_info['NAME']}</b>", company_style),
            Paragraph(f"{company_info['LOCATION']}", company_style),
            Paragraph(f"Email: {company_info['EMAIL']} | Tel: {company_info['PHONE']}", company_style),
            Spacer(1, 0.5 * cm),
            Paragraph('Saldos', styles['Title']),
            Spacer(1, 0.5 * cm),
        ]
        available_width = landscape(A4)[0] - 3 * cm
        col_width = available_width / len(headers)
        table_data = [headers] + rows + [totals]
        table = Table(table_data, colWidths=[col_width] * len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        path = f'exports/{self.request.id}_{filename}'
        buffer.seek(0)
        default_storage.save(path, buffer)

    logger.info('Export async concluído: %s', path)
    _notify(user_email, f'Export Saldos ({export_format.upper()})', path)
    return {'status': 'ok', 'path': path}
